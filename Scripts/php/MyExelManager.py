from openpyxl import load_workbook

filemodes = ['r','w']

currentfm = 0

nums = [1,2,3,4,5,6,7,8,9]

currentnum = 0

#main programm cycle
start = True
while start:

            action = input('action : ') #READ FILE

            row = int(input('row : '))

            letter = input('letter : ')
    
            if action =='read file':

                filename = input('nazvanie faila : ')

                wb = load_workbook(str(filename)) # Load in the workbook #wb -> workbook

                filemode = filemodes[currentfm]
                
                file = open(str(filename),filemode)
                
                wsactive = wb.active #ws -> worksheet #wb -> workbook  

                print()
                for i in range (len(nums)):
                    print(wsactive[letter + str(nums[i])].value)

                print()                    
                
                file.close()

                wb.close()

            if action =='write file':#WRITE FILE

                filemode = filemodes[currentfm]

                filename = input('nazvanie faila : ')

                myvalue = input('value : ')

                f = open(str(filename),filemode)

                wb = load_workbook(str(filename))

                ws = wb.active

                ws[letter + str(nums[row])].value = myvalue
                    
                f.close()

                wb.save(str(filename))
                
                wb.close()










"""            #set a file mode 'w'- write in file  , 'r'- read file 
            if action =='write file':
                filemode = 'w'
                f = open(str(filename),filemode)

                # Get sheet names
                #writing values from results in exel file

                for i in range(1,len(Results)):
                    sheet['A'+str(i)].value =str(Results[i])
                    print(sheet['A' + str(i)].value)
                f.close()

"""
                